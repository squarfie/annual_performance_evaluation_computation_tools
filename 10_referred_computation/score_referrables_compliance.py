from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =====================================================
# SETTINGS
# =====================================================
NON_REQUIRED_BASELINE = 3

BREAKPOINT_ORGANISM_ALIASES = {
    "eco": ["ebc"],
    "kpn": ["ebc"],
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_or(value: Any) -> set[str]:
    return {
        part.strip().lower()
        for part in clean(value).split("|")
        if part.strip()
    }


def parse_antibiotic_expression(value: Any) -> list[list[str]]:
    text = clean(value)

    if not text:
        return []

    groups = []

    if "&" in text:
        for and_part in text.split("&"):
            alternatives = [
                part.strip().upper()
                for part in and_part.split("|")
                if part.strip()
            ]

            if alternatives:
                groups.append(alternatives)

    else:
        alternatives = [
            part.strip().upper()
            for part in text.split("|")
            if part.strip()
        ]

        if alternatives:
            groups.append(alternatives)

    return groups


def normalize_test_aliases(test_code: str) -> list[str]:
    test_code = clean(test_code).upper()

    aliases = [test_code]

    if "_ND30" in test_code:
        aliases.append(test_code.replace("_ND30", "_ND10"))

    if "_ND10" in test_code:
        aliases.append(test_code.replace("_ND10", "_ND30"))

    return list(dict.fromkeys(aliases))


def parse_result_number(value: Any) -> float | None:
    text = clean(value).replace(" ", "")

    if not text:
        return None

    text = (
        text.replace("<=", "")
        .replace(">=", "")
        .replace("<", "")
        .replace(">", "")
        .replace(",", ".")
    )

    match = re.search(r"\d+(?:\.\d+)?", text)

    return float(match.group(0)) if match else None


def parse_breakpoint_range(value: Any) -> tuple[float, float] | None:
    text = clean(value).replace(" ", "")

    if not text:
        return None

    nums = [
        float(x)
        for x in re.findall(r"\d+(?:\.\d+)?", text)
    ]

    if not nums:
        return None

    if len(nums) == 1:
        return nums[0], nums[0]

    return min(nums[0], nums[1]), max(nums[0], nums[1])


def classify_result(
    value: Any,
    breakpoint: dict[str, Any] | None,
    test_code: str
) -> str | None:

    raw = clean(value).upper()

    if raw in {"S", "I", "R", "NS", "SDD"}:
        return raw

    if breakpoint is None:
        return None

    result = parse_result_number(value)

    if result is None:
        return None

    is_mic = (
        "_NM" in test_code.upper()
        or "_NE" in test_code.upper()
    )

    s_bp = parse_breakpoint_range(breakpoint.get("S"))
    i_bp = parse_breakpoint_range(breakpoint.get("I"))
    sdd_bp = parse_breakpoint_range(breakpoint.get("SDD"))
    r_bp = parse_breakpoint_range(breakpoint.get("R"))

    if is_mic:
        if s_bp and result <= s_bp[1]:
            return "S"

        if i_bp and i_bp[0] <= result <= i_bp[1]:
            return "I"

        if sdd_bp and sdd_bp[0] <= result <= sdd_bp[1]:
            return "SDD"

        if r_bp and result >= r_bp[0]:
            return "R"

    else:
        if s_bp and result >= s_bp[0]:
            return "S"

        if i_bp and i_bp[0] <= result <= i_bp[1]:
            return "I"

        if sdd_bp and sdd_bp[0] <= result <= sdd_bp[1]:
            return "SDD"

        if r_bp and result <= r_bp[1]:
            return "R"

    return None


def get_breakpoint(
    breakpoints: dict[tuple[str, str], dict[str, Any]],
    organism: str,
    test_code: str,
) -> dict[str, Any] | None:

    organism_key = clean(organism).lower()
    test_key = clean(test_code).upper()

    for candidate in [
        organism_key,
        *BREAKPOINT_ORGANISM_ALIASES.get(organism_key, [])
    ]:
        breakpoint = breakpoints.get((candidate, test_key))

        if breakpoint is not None:
            return breakpoint

    return None


def parse_age_to_days(value: Any) -> int | None:
    text = clean(value).lower()

    if not text:
        return None

    if text in {"0", "0.0"}:
        return 0

    match = re.fullmatch(r"(\d+)\s*d", text)

    if match:
        return int(match.group(1))

    match = re.fullmatch(r"(\d+)\s*m", text)

    if match:
        return int(match.group(1)) * 30

    match = re.fullmatch(r"(\d+)\s*y", text)

    if match:
        return int(match.group(1)) * 365

    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        return int(float(text) * 365)

    return None


def age_matches(patient_age: Any, criterion_age: Any) -> bool:
    crit = clean(criterion_age).lower()

    if not crit:
        return True

    patient_days = parse_age_to_days(patient_age)

    if patient_days is None:
        return False

    if crit == "<1":
        return 0 <= patient_days <= 28

    match = re.fullmatch(
        r"([<>]=?|=)?\s*(\d+(?:\.\d+)?)",
        crit
    )

    if not match:
        return clean(patient_age).lower() == crit

    op = match.group(1) or "="

    crit_days = int(float(match.group(2)) * 365)

    if op == "<":
        return patient_days < crit_days

    if op == "<=":
        return patient_days <= crit_days

    if op == ">":
        return patient_days > crit_days

    if op == ">=":
        return patient_days >= crit_days

    return patient_days == crit_days


def rule_day_target(rule: dict[str, Any]) -> int | None:
    if clean(rule.get("days")) == "":
        return None

    try:
        target = int(float(clean(rule["days"])))
    except ValueError:
        return None

    return target if target > 0 else None


def load_referrables(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    headers = [
        clean(v)
        for v in next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True
            )
        )
    ]

    header_index = {
        name: idx
        for idx, name in enumerate(headers)
        if name
    }

    required = [
        "Organism_Code",
        "Specimen_code",
        "Antibiotic_code",
        "Days",
        "Interpretation",
        "Age"
    ]

    missing = [
        name
        for name in required
        if name not in header_index
    ]

    if missing:
        raise ValueError(
            f"Referrables list is missing columns: "
            f"{', '.join(missing)}"
        )

    rules = []

    for excel_row, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):

        organism = clean(row[header_index["Organism_Code"]])
        specimen = clean(row[header_index["Specimen_code"]])

        if not organism or not specimen:
            continue

        rule = {
            "rule_id": len(rules) + 1,
            "source_row": excel_row,
            "organisms": split_or(organism),
            "specimens": split_or(specimen),
            "antibiotic_expression": clean(
                row[header_index["Antibiotic_code"]]
            ),
            "antibiotic_groups": parse_antibiotic_expression(
                row[header_index["Antibiotic_code"]]
            ),
            "days": row[header_index["Days"]],
            "interpretations": split_or(
                row[header_index["Interpretation"]]
            ),
            "age": row[header_index["Age"]],
            "raw_organism": organism,
            "raw_specimen": specimen,
        }

        rules.append(rule)

    wb.close()

    return rules


def load_breakpoints(path: Path) -> dict[tuple[str, str], dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Breakpoints"]

    headers = [
        clean(v)
        for v in next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True
            )
        )
    ]

    idx = {
        h: i
        for i, h in enumerate(headers)
    }

    needed = [
        "ORGANISM_CODE",
        "WHONET_TEST",
        "R",
        "I",
        "SDD",
        "S"
    ]

    missing = [
        name
        for name in needed
        if name not in idx
    ]

    if missing:
        raise ValueError(
            f"Breakpoints file is missing columns: "
            f"{', '.join(missing)}"
        )

    breakpoints = {}

    for row in ws.iter_rows(min_row=2, values_only=True):

        organism = clean(row[idx["ORGANISM_CODE"]]).lower()
        test = clean(row[idx["WHONET_TEST"]]).upper()

        if not organism or not test:
            continue

        bp = {
            name: row[idx[name]]
            for name in ["R", "I", "SDD", "S"]
        }

        if not any(clean(v) for v in bp.values()):
            continue

        breakpoints.setdefault((organism, test), bp)

    wb.close()

    return breakpoints


def find_column(
    headers: dict[str, int],
    requested_test: str
) -> tuple[str, int] | tuple[None, None]:

    for alias in normalize_test_aliases(requested_test):

        if alias in headers:
            return alias, headers[alias]

    return None, None


def antibiotic_alternative_matches(
    row: tuple[Any, ...],
    headers: dict[str, int],
    organism: str,
    alternatives: list[str],
    allowed_interpretations: set[str],
    breakpoints: dict[tuple[str, str], dict[str, Any]],
) -> tuple[bool, list[str]]:

    notes = []

    for requested_test in alternatives:

        actual_test, col_idx = find_column(
            headers,
            requested_test
        )

        if actual_test is None or col_idx is None:
            notes.append(f"{requested_test}: missing column")
            continue

        result_value = row[col_idx]

        if clean(result_value) == "":
            notes.append(f"{requested_test}: blank")
            continue

        if allowed_interpretations:

            bp = get_breakpoint(
                breakpoints,
                organism,
                actual_test
            )

            interpretation = classify_result(
                result_value,
                bp,
                actual_test
            )

            if interpretation is None:
                notes.append(
                    f"{requested_test}: cannot classify "
                    f"{clean(result_value)}"
                )
                continue

            if interpretation.lower() not in allowed_interpretations:
                notes.append(
                    f"{requested_test}: "
                    f"{clean(result_value)} => {interpretation}"
                )
                continue

            notes.append(
                f"{requested_test}: "
                f"{clean(result_value)} => {interpretation}"
            )

            return True, notes

        else:

            notes.append(
                f"{requested_test}: present "
                f"{clean(result_value)}"
            )

            return True, notes

    return False, notes


def rule_matches_row(
    rule: dict[str, Any],
    row: tuple[Any, ...],
    headers: dict[str, int],
    breakpoints: dict[tuple[str, str], dict[str, Any]],
) -> tuple[bool, str]:

    organism = clean(row[headers["ORGANISM"]]).lower()
    specimen = clean(row[headers["SPEC_TYPE"]]).lower()

    if organism not in rule["organisms"]:
        return False, ""

    if specimen not in rule["specimens"]:
        return False, ""

    if not age_matches(row[headers["AGE"]], rule["age"]):
        return False, ""

    and_groups = rule["antibiotic_groups"]

    if not and_groups:
        return True, "Organism/specimen matched; no antibiotic criterion"

    all_notes = []

    for alternatives in and_groups:

        ok, notes = antibiotic_alternative_matches(
            row,
            headers,
            organism,
            alternatives,
            rule["interpretations"],
            breakpoints,
        )

        all_notes.extend(notes)

        if not ok:
            return False, ""

    return True, "; ".join(all_notes)


def get_sortable_date(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value

    text = clean(value)

    for fmt in (
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%Y/%m/%d"
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass

    return datetime.max


def score_consolidated_file(
    path: Path,
    rules: list[dict[str, Any]],
    breakpoints: dict[tuple[str, str], dict[str, Any]]
) -> list[dict[str, Any]]:

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers_list = [
        clean(v)
        for v in next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True
            )
        )
    ]

    headers = {
        name: idx
        for idx, name in enumerate(headers_list)
        if name
    }

    required = [
        "LABORATORY",
        "PATIENT_ID",
        "SPEC_NUM",
        "ORGANISM",
        "SPEC_TYPE",
        "SPEC_DATE",
        "AGE"
    ]

    missing = [
        name
        for name in required
        if name not in headers
    ]

    if missing:
        raise ValueError(
            f"{path.name} is missing columns: "
            f"{', '.join(missing)}"
        )

    required_rules = [
        rule
        for rule in rules
        if clean(rule.get("days")) != ""
    ]

    non_required_rules = [
        rule
        for rule in rules
        if clean(rule.get("days")) == ""
    ]

    site_matches = defaultdict(lambda: defaultdict(list))
    site_match_counts = defaultdict(lambda: defaultdict(int))
    site_totals = defaultdict(int)

    for row_number, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):

        laboratory = (
            clean(row[headers["LABORATORY"]])
            or "UNKNOWN"
        )

        site_totals[laboratory] += 1

        for rule in rules:

            ok, note = rule_matches_row(
                rule,
                row,
                headers,
                breakpoints
            )

            if ok:

                match_data = {
                    "file": path.name,
                    "laboratory": laboratory,
                    "rule_id": rule["rule_id"],
                    "rule_row": rule["source_row"],
                    "source_data_row": row_number,
                    "patient_id": clean(row[headers["PATIENT_ID"]]),
                    "spec_num": clean(row[headers["SPEC_NUM"]]),
                    "organism": clean(row[headers["ORGANISM"]]),
                    "specimen": clean(row[headers["SPEC_TYPE"]]),
                    "spec_date": row[headers["SPEC_DATE"]],
                    "age": clean(row[headers["AGE"]]),
                    "note": note,
                }

                site_matches[laboratory][rule["rule_id"]].append(
                    match_data
                )

                site_match_counts[laboratory][rule["rule_id"]] += 1

    wb.close()

    scored_sites = []

    for laboratory in sorted(site_totals):

        matched_by_rule = site_matches[laboratory]
        matched_counts = site_match_counts[laboratory]

        required_points = 0
        required_complied = 0

        for rule in required_rules:

            target = rule_day_target(rule)

            all_matches = matched_by_rule.get(
                rule["rule_id"],
                []
            )

            all_matches = sorted(
                all_matches,
                key=lambda x: (
                    get_sortable_date(x["spec_date"]),
                    x["source_data_row"]
                )
            )

            counted_matches = (
                all_matches[:target]
                if target
                else []
            )

            counted_count = len(counted_matches)

            if target and target > 0:
                criterion_score = min(
                    counted_count / target,
                    1
                )
            else:
                criterion_score = 0

            if criterion_score >= 1:
                required_complied += 1

            required_points += criterion_score

        if required_rules:
            required_percent = (
                required_points
                / len(required_rules)
            ) * 100
        else:
            required_percent = 0

        non_required_complied = 0

        for rule in non_required_rules:

            matches = matched_counts.get(
                rule["rule_id"],
                0
            )

            if matches > 0:
                non_required_complied += 1

        if non_required_complied == 0:
            non_required_percent = 0

        elif non_required_complied < NON_REQUIRED_BASELINE:
            non_required_percent = 50

        else:
            non_required_percent = 100

        final_score = (
            required_percent
            + non_required_percent
        ) / 2

        scored_sites.append(
            {
                "file": path.name,
                "laboratory": laboratory,
                "total_records": site_totals[laboratory],

                "required_rules": len(required_rules),
                "required_complied": required_complied,
                "required_score": round(required_percent, 2),

                "non_required_rules": len(non_required_rules),
                "non_required_complied": non_required_complied,
                "non_required_baseline": NON_REQUIRED_BASELINE,
                "non_required_score": round(non_required_percent, 2),

                "final_score": round(final_score, 2),

                "matched_by_rule": matched_by_rule,
                "matched_counts": matched_counts,
            }
        )

    return scored_sites


def autosize(ws) -> None:
    for col_idx, column_cells in enumerate(
        ws.columns,
        start=1
    ):

        max_len = 0

        for cell in column_cells:
            max_len = max(max_len, len(clean(cell.value)))

        ws.column_dimensions[
            get_column_letter(col_idx)
        ].width = min(max(max_len + 2, 10), 60)


def safe_sheet_name(name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\[\]\*\?/\\:]", "_", name).strip() or "Sheet"

    cleaned = cleaned[:31]

    candidate = cleaned

    counter = 2

    while candidate in used_names:

        suffix = f"_{counter}"

        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"

        counter += 1

    used_names.add(candidate)

    return candidate


def criterion_sheet_label(rule: dict[str, Any]) -> str:
    organism = clean(
        rule["raw_organism"]
    ).split("|", 1)[0].strip().lower()

    row_label_overrides = {
        7: "sal",
        8: "sal",
        9: "shi",
    }

    return row_label_overrides.get(
        rule["source_row"],
        organism
    )


def append_match_row(ws, match: dict[str, Any]) -> None:
    ws.append(
        [
            match["file"],
            match["laboratory"],
            match["rule_id"],
            match["rule_row"],
            match["source_data_row"],
            match["patient_id"],
            match["spec_num"],
            match["organism"],
            match["specimen"],
            match["spec_date"],
            match["age"],
            match["note"],
        ]
    )


def add_organism_match_header(ws) -> None:
    ws.append(
        [
            "Rule ID",
            "Referrables Row",
            "Criteria Organism_Code",
            "Criteria Specimen_code",
            "Criteria Antibiotic_code",
            "Criteria Days / First N",
            "Criteria Interpretation",
            "Criteria Age",
            "File",
            "Laboratory",
            "Data Row",
            "Patient ID",
            "Spec Num",
            "Organism",
            "Specimen",
            "Spec Date",
            "Age",
            "Antibiotic/Interpretation Note",
        ]
    )


def append_organism_match_row(
    ws,
    rule: dict[str, Any],
    match: dict[str, Any]
) -> None:

    ws.append(
        [
            rule["rule_id"],
            rule["source_row"],
            rule["raw_organism"],
            rule["raw_specimen"],
            rule["antibiotic_expression"],
            rule["days"],
            "|".join(sorted(rule["interpretations"])),
            rule["age"],
            match["file"],
            match["laboratory"],
            match["source_data_row"],
            match["patient_id"],
            match["spec_num"],
            match["organism"],
            match["specimen"],
            match["spec_date"],
            match["age"],
            match["note"],
        ]
    )


def write_report(
    output_path: Path,
    rules: list[dict[str, Any]],
    scored_sites: list[dict[str, Any]]
) -> None:

    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"

    summary.append(
        [
            "Sentinel Site",
            "Source File",
            "Total Records",
            "Required Rules",
            "Required Complied",
            "Required Score %",
            "Non-Required Rules",
            "Non-Required Complied",
            "Non-Required Baseline",
            "Non-Required Score %",
            "Final Annual Score %",
        ]
    )

    for site in scored_sites:

        summary.append(
            [
                site["laboratory"],
                site["file"],
                site["total_records"],
                site["required_rules"],
                site["required_complied"],
                site["required_score"],
                site["non_required_rules"],
                site["non_required_complied"],
                site["non_required_baseline"],
                site["non_required_score"],
                site["final_score"],
            ]
        )

    criteria = wb.create_sheet("Criteria_Status")

    criteria.append(
        [
            "Sentinel Site",
            "Rule Type",
            "Rule ID",
            "Referrables Row",
            "Organism_Code",
            "Specimen_code",
            "Antibiotic_code",
            "Days / First N",
            "Interpretation",
            "Age",
            "Status",
            "Criterion Score %",
            "Total Match Count",
            "Counted Match Count",
        ]
    )

    counts_ws = wb.create_sheet("Criteria_Isolate_Counts")

    counts_ws.append(
        [
            "Sentinel Site",
            "Rule Type",
            "Rule ID",
            "Referrables Row",
            "Organism_Code",
            "Specimen_code",
            "Antibiotic_code",
            "Days / First N",
            "Interpretation",
            "Age",
            "Total Match Count",
            "Counted Match Count",
            "Criterion Score %",
        ]
    )

    for site in scored_sites:

        for rule in rules:

            all_matches = site["matched_by_rule"].get(
                rule["rule_id"],
                []
            )

            total_match_count = site["matched_counts"].get(
                rule["rule_id"],
                0
            )

            if clean(rule.get("days")) != "":

                rule_type = "Required / First N"

                target = rule_day_target(rule)

                sorted_matches = sorted(
                    all_matches,
                    key=lambda x: (
                        get_sortable_date(x["spec_date"]),
                        x["source_data_row"]
                    )
                )

                counted_matches = (
                    sorted_matches[:target]
                    if target
                    else []
                )

                counted_match_count = len(counted_matches)

                criterion_score = (
                    min(counted_match_count / target, 1) * 100
                    if target
                    else 0
                )

                if criterion_score >= 100:
                    status = "YES"
                elif criterion_score > 0:
                    status = "PARTIAL"
                else:
                    status = "NO"

            else:

                rule_type = "Non-Required / Presence"

                counted_match_count = total_match_count

                criterion_score = (
                    100
                    if total_match_count > 0
                    else 0
                )

                status = (
                    "YES"
                    if total_match_count > 0
                    else "NO"
                )

            criteria.append(
                [
                    site["laboratory"],
                    rule_type,
                    rule["rule_id"],
                    rule["source_row"],
                    rule["raw_organism"],
                    rule["raw_specimen"],
                    rule["antibiotic_expression"],
                    rule["days"],
                    "|".join(sorted(rule["interpretations"])),
                    rule["age"],
                    status,
                    round(criterion_score, 2),
                    total_match_count,
                    counted_match_count,
                ]
            )

            counts_ws.append(
                [
                    site["laboratory"],
                    rule_type,
                    rule["rule_id"],
                    rule["source_row"],
                    rule["raw_organism"],
                    rule["raw_specimen"],
                    rule["antibiotic_expression"],
                    rule["days"],
                    "|".join(sorted(rule["interpretations"])),
                    rule["age"],
                    total_match_count,
                    counted_match_count,
                    round(criterion_score, 2),
                ]
            )

    details = wb.create_sheet("Matched_Records")

    details.append(
        [
            "File",
            "Laboratory",
            "Rule ID",
            "Referrables Row",
            "Data Row",
            "Patient ID",
            "Spec Num",
            "Organism",
            "Specimen",
            "Spec Date",
            "Age",
            "Antibiotic/Interpretation Note",
        ]
    )

    for site in scored_sites:

        for matches in site["matched_by_rule"].values():

            for match in matches:
                append_match_row(details, match)

    rules_by_label = defaultdict(list)

    for rule in rules:
        rules_by_label[criterion_sheet_label(rule)].append(rule)

    used_sheet_names = {
        ws.title
        for ws in wb.worksheets
    }

    for label, label_rules in rules_by_label.items():

        sheet_name = safe_sheet_name(
            label,
            used_sheet_names
        )

        ws = wb.create_sheet(sheet_name)

        add_organism_match_header(ws)

        for rule in label_rules:

            for site in scored_sites:

                for match in site["matched_by_rule"].get(
                    rule["rule_id"],
                    []
                ):

                    append_organism_match_row(
                        ws,
                        rule,
                        match
                    )

    for ws in wb.worksheets:

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = PatternFill(
                "solid",
                fgColor="1F4E78"
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        for row in ws.iter_rows():

            for cell in row:

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

        autosize(ws)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    wb.save(output_path)


def main() -> None:

    from tkinter import Tk, filedialog

    Tk().withdraw()

    print("\n===================================")
    print("REFERRED ISOLATE COMPLIANCE SCORER")
    print("===================================\n")

    referrables_path = filedialog.askopenfilename(
        title="Select Referrables_list.xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )

    if not referrables_path:
        print("No Referrables file selected.")
        return

    breakpoints_path = filedialog.askopenfilename(
        title="Select Breakpoints.xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )

    if not breakpoints_path:
        print("No Breakpoints file selected.")
        return

    consolidated_path = filedialog.askopenfilename(
        title="Select Consolidated WHONET Excel File",
        filetypes=[("Excel Files", "*.xlsx")]
    )

    if not consolidated_path:
        print("No consolidated file selected.")
        return

    output_path = filedialog.asksaveasfilename(
        title="Save Compliance Report As",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")],
        initialfile="referrables_compliance_report.xlsx"
    )

    if not output_path:
        print("No output file selected.")
        return

    referrables_path = Path(referrables_path)
    breakpoints_path = Path(breakpoints_path)
    consolidated_path = Path(consolidated_path)
    output_path = Path(output_path)

    print("\nLoading rules...")
    rules = load_referrables(referrables_path)

    print("Loading breakpoints...")
    breakpoints = load_breakpoints(breakpoints_path)

    print("Scoring sentinel sites...")
    scored_sites = score_consolidated_file(
        consolidated_path,
        rules,
        breakpoints
    )

    print("Generating report...")
    write_report(
        output_path,
        rules,
        scored_sites
    )

    print("\n===================================")
    print("✔ REPORT GENERATED SUCCESSFULLY")
    print("===================================")
    print(f"Output File: {output_path}")
    print(f"Sites Scored: {len(scored_sites)}")
    print(f"Criteria Loaded: {len(rules)}")
    print(f"Non-required baseline: {NON_REQUIRED_BASELINE}")


if __name__ == "__main__":
    main()